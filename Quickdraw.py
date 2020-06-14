from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import sys
import webbrowser

# Function for getting user response, since people find a way to input the strangest things
# Either that, or they're trying to intentionally break the program >.>
def get_from_user(message, valid_resp = [], content_type = "string"):
    while True:
        user_resp = input(message).lower()
        if content_type == "int":
            try:
                int(user_resp)
            except ValueError:
                continue
            break
        elif content_type == "file":
            file_path = Path(user_resp)
            if file_path.is_file():
                break
        elif len(valid_resp) > 0:
            if user_resp in valid_resp:
                break
        else:
            break
    return user_resp

# Function to format and find the time difference between two Spyglass time strings
def time_diff(time_string1, time_string2):
    dateobj1 = datetime.strptime(time_string1, "%H:%M:%S")
    dateobj2 = datetime.strptime(time_string2, "%H:%M:%S")
    return int(abs((dateobj1 - dateobj2).total_seconds()))

# Function to strip a region of any Spyglass indicators at the end of it's name (* or ~)
def sanitize_region(region):
    sanitized_region = region
    if region[-1] == "~" or region[-1] == "*":
        sanitized_region = sanitized_region[:-1]
    sanitized_region = sanitized_region.lower()
    return sanitized_region.replace(" ", "_")

# Function to write output to the raid_file and trigger_list, given the target, trigger, and target number
def write_to_outfile(targ_number, targ, trigger, targ_upd_time, trigger_len):
    with open("raid_file.txt", "a") as file:
        file.write(f"{targ_number}) https://www.nationstates.net/region={sanitize_region(targ)} ({targ_upd_time})\n")
        file.write(f"    a) https://www.nationstates.net/template-overall=none/region={sanitize_region(trigger)} ({trigger_len}s)\n\n")

    with open("trigger_list.txt", "a") as file:
        file.write(f"{trigger}\n")

# Create a class to hold target info, since I'd prefer to keep the target_array as one list
# Should hold region name, update time, and row on the Spyglass sheet
class Target:
    def __init__(self, name, update_time, row):
        self.name = name
        self.update_time = update_time
        self.row = row

# Create a class for triggers
class Trigger:
    def __init__(self, name, score, length):
        self.name = name
        self.score = score
        self.length = length

# Class for switches
class Switch:
    def __init__(self, name, pos):
        self.name = name
        self.pos = pos

# Create or overwrite any already existing raid_file.txt or trigger_list.txt files
with open("raid_file.txt", "w") as file:
    file.write("")

with open("trigger_list.txt", "w") as file:
    file.write("")

# Grab the various parameters we need to start finding targets
sheet_path = get_from_user("Enter path to a Spyglass sheet: ", content_type = "file")
targ_amount = int(get_from_user("Amount of targets: ", content_type = "int"))
min_switch_length = int(get_from_user("Minimum switch length: ", content_type = "int"))
min_trigger_length = int(get_from_user("Minimum trigger length: ", content_type = "int"))
optim_trigger_length = int(get_from_user("Optimum trigger length: ", content_type = "int"))
max_trigger_length = int(get_from_user("Maximum trigger length: ", content_type = "int"))
endos = int(get_from_user("Endorsements on point: ", content_type = "int"))
embassy_filters = get_from_user("Ignore embassies: ")
wfe_filters = get_from_user("Ignore phrases: ")
update = get_from_user("Major or Minor update: ", ["major", "minor"])

# Set the appropriate columns to grab times from on the Spyglass sheet, depending on which update is selected
if update == "major":
    time_column = 6
else:
    time_column = 5

# Load the Spyglass sheet specified by the user so that we can start reading off of it
# We can now read cells of the sheet from ws
wb = load_workbook(sheet_path, read_only = True)
ws = wb.active
sheet_length = ws.max_row

# Now that we have all the user input and the sheet loaded, we can finish writing functions that work off of that
# Even though we're filling in the target_array later, I'll define it here, since I want to vaguely keep my functions defined in the same place
target_array = []

# Function to find a trigger for a region given a row of a Spyglass sheet for the target
# Will either return a ConstraintRegion object, or a an empty string if no trigger is found
def find_trigger(spyglass_row):
    trigger_list = []
    i = 1
    base_region_update = ws.cell(row = spyglass_row, column = time_column).value
    trigger_update = ws.cell(row = spyglass_row - i, column = time_column).value

    # Put this in a try/except block in case it goes back too far looking for triggers, and tries to parse the headers
    # If that happens, obviously no trigger was found, so we can just return an empty string
    try:
        trigger_length = time_diff(base_region_update, trigger_update)
    except ValueError:
        return ""

    while trigger_length <= max_trigger_length:
        if trigger_length >= min_trigger_length:
            trigger_name = ws.cell(row = spyglass_row - i, column = 1).value
            trigger_score = abs(trigger_length - optim_trigger_length)

            trigger_list.append(Trigger(trigger_name, trigger_score, trigger_length))

        i += 1
        trigger_update = ws.cell(row = spyglass_row - i, column = time_column).value
        try:
            trigger_length = time_diff(base_region_update, trigger_update)
        except ValueError:
            return ""

    trigger_list.sort(key = lambda region: region.score)
    try:
        return trigger_list[0]
    except IndexError:
        return ""

# Function to find a switch given a region's position in the target_array
# Will return the target_array index of the switch closest to the minimum switch length specified
def find_switch(target_array_pos):
    i = 1
    base_region_update = target_array[target_array_pos].update_time
    switch_update = target_array[target_array_pos + i].update_time
    while time_diff(base_region_update, switch_update) < min_switch_length:
        i += 1
        switch_update = target_array[target_array_pos + i].update_time
    return target_array_pos + i

# Split the embassy and WFE filter strings into nice arrays to make filtering easier later
# Also, make sure to trim whitespace, since that will affect filtering
embassy_filter_list = list(map(lambda filter: filter.strip(), embassy_filters.split(",")))
wfe_filter_list = list(map(lambda filter: filter.strip(), wfe_filters.split(",")))

# If a user leaves the filters empty, we can be reasonably sure that they don't care about retagging
# However, even if no input is given, the list will still have one empty string in it, which will cause *every* region to be filtered out
# To get around this, we will populate the first item with a long random string, which will not filter out any regions
if embassy_filter_list[0] == "":
    embassy_filter_list[0] = "eijl3o2il21po21-0p1ojiqlksakox;l"
if wfe_filter_list[0] == "":
    wfe_filter_list[0] = "OIWJQLKDWLK<WJQPOLKJQWLlqwwq0qwqwq0wqwp1kl"

# Okay, now we can finally get to filling the target_array like I said I would do earlier
# Basically, we just check if there's a * at the end of a target to determine if it's hittable, and if it is, we add it to the target_array
# Oh, and also make sure it doesn't get filtered out by endo count or embassies or wfe, that'd be good too
print("Generating list of valid targets...")
row_counter = 2
for row in ws.iter_rows(min_row = 2, max_col = 10):
    region_name = row[0].value
    region_del_endos = int(row[7].value)

    if region_name[-1] == "~" and region_del_endos < endos:

        # Put the parsing of region embassies and WFE into a try/except block, in case it's empty on the Spyglass sheet
        try:
            region_embassies = row[8].value.lower()
        except AttributeError:
            region_embassies = ""

        for filter in embassy_filter_list:
            if filter in region_embassies:
                break
        else:
            try:
                region_wfe = row[9].value.lower()
            except AttributeError:
                region_wfe = ""

            for filter in wfe_filter_list:
                if filter in region_wfe:
                    break
            else:
                region_update = row[time_column - 1].value
                target_array.append(Target(region_name, region_update, row_counter))

    row_counter += 1

# Now that we have generated an array of targets, we can start finding switches and triggers
# We should use seperate variables to iterate over the target_array and to count the actual targets, since we'll be jumping around a lot in the target array to get switches
target_counter = 1
target_array_pos = 0

# Keep going through the target_array list until we either have enough targs, or run past the end of the list
while target_array_pos < len(target_array) and target_counter <= targ_amount:
    current_target = target_array[target_array_pos]
    current_trigger = find_trigger(current_target.row)

    # We're basically checking to see if a trigger was found, since the find_trigger method will return an empty string if no triggers are found
    if current_trigger != "":
        target_name = current_target.name
        target_update = current_target.update_time
        target_url = f"https://www.nationstates.net/region={sanitize_region(target_name)}"
        webbrowser.open_new_tab(target_url)
        valid_targ = get_from_user(f"{target_counter}) Is {target_name[:-1]} ({target_update}) an acceptable target? (y/n) ", ["y", "n"])

        if valid_targ == "y":
            trigger_name = current_trigger.name
            trigger_length = current_trigger.length
            write_to_outfile(target_counter, target_name, trigger_name, target_update, trigger_length)

            # Now we just find the next switch after minimum switch length, and repeat the process
            # Put in a try/except block, since running past the end of our target_array list will cause an index error
            try:
                # Subtract one, since the target_array iterator will increment at the end of the loop regardless
                target_array_pos = find_switch(target_array_pos) - 1
            except IndexError:
                print(f"Reached end of update! Found {target_counter}/{targ_amount} targets")
                sys.exit()

            target_counter += 1

    target_array_pos += 1

print("Finished finding all targets!")
sys.exit()
